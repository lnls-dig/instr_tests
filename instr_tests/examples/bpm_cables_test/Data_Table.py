#! /usr/bin/env python

# This script is responsible for searching specific data files on cable tests, and then creating an organized table with all the desired information.
# It complements the testing script 'cable_test.py' found at same repository as this one.
# Leticia Garcez Capovilla - LNLS

import sys
import os
import fnmatch
import openpyxl
import openpyxl.styles
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# path to current directory (script must be in the same directory as data folders)
path = os.getcwd()

# lists of data categories
matches = []
manufac = []
size = []
label = []
color = []

# column titles
c_a = 'Manufacturer'
c_b = 'Size'
c_c = 'Label'
c_d = 'Color'
c_e = 'S11@ 1MHz(red)'
c_f = 'S11@ 500MHz'
c_g = 'S11@ AVG'
c_h = 'S21@ 1MHz(red)'
c_i = 'S21@ 500MHz'
c_j = 'S21@ 3GHz'
c_k = 'S22@ 1MHz(red)'
c_l = 'S22@ 500MHz'
c_m = 'S22@ AVG'
c_n = 'Delay'
c_o = 'Quality of crimping / thermo-shrinkable / robustness (Excellent / Good / Bad)'
c_p = 'Quality of Identification'
c_q = 'Observation'


# walk down the directory and find all statistics files
for root, dirs, fnames in os.walk(path):
	for fname in fnmatch.filter(fnames, '*statistics_*'):
		matches.append(os.path.join(root, fname))


# create lists of categories: Manufacturers, Sizes, Labels and Colors of the tested cables
for m in matches:
	man = m.rsplit('/',5)[1]										# get the manufacturer directory
	manufac.append(man)											# make a list of the names
	siz = m.rsplit('/',4)[1]
	size.append(siz)
	lab = m.rsplit('/',3)[1]
	label.append(lab)
	col = m.rsplit('/',2)[1]
	color.append(col)


# create csv file with all the statistics data 
with open('Data.csv','w') as dtfile:
	for m in matches:
		rfile = open(m)
		lines = rfile.readlines()
		for line in lines:
			elem = line.replace('\n','').split('\t')[1]
			dtfile.write(elem+'\t')
		dtfile.write('\n')


# create csv file with all inspections data
i=0
with open('Inspection.csv','w') as inspfile:
	while i <= (len(manufac)-1):
		rfile = open(path+'/'+manufac[i]+'/'+size[i]+'/'+label[i]+'/inspection_'+label[i]+'.txt')
		lines = rfile.readlines()
		for line in lines:
			elem = line.replace('\n','').split('\t')[1]
			inspfile.write(elem+'\t')
		inspfile.write('\n')
		i = i+1


# create workbook
wb = Workbook()

# activate default worksheet
ws = wb.active

# define worksheet title
ws.title = 'Cable_tests'

# define reference positions
head_line = 1

# define reference format
head_font = Font(bold=True)											# makes the font bold
default_align = Alignment(horizontal = 'center', vertical = 'center')						# align the words in the center
df_wid = 20

# header
ws['A'+str(head_line)] = c_a											# what is written in the cell
ws['A'+str(head_line)].font = head_font
ws['A'+str(head_line)].alignment = default_align
ws.column_dimensions['A'].width = 15

ws['B'+str(head_line)] = c_b
ws['B'+str(head_line)].font = head_font
ws['B'+str(head_line)].alignment = default_align
ws.column_dimensions['B'].width = 10

ws['C'+str(head_line)] = c_c
ws['C'+str(head_line)].font = head_font
ws['C'+str(head_line)].alignment = default_align
ws.column_dimensions['C'].width = 15

ws['D'+str(head_line)] = c_d
ws['D'+str(head_line)].font = head_font
ws['D'+str(head_line)].alignment = default_align
ws.column_dimensions['D'].width = 10

ws['E'+str(head_line)] = c_e
ws['E'+str(head_line)].font = head_font
ws['E'+str(head_line)].alignment = default_align
ws.column_dimensions['E'].width = df_wid

ws['F'+str(head_line)] = c_f
ws['F'+str(head_line)].font = head_font
ws['F'+str(head_line)].alignment = default_align
ws.column_dimensions['F'].width = df_wid

ws['G'+str(head_line)] = c_g
ws['G'+str(head_line)].font = head_font
ws['G'+str(head_line)].alignment = default_align
ws.column_dimensions['G'].width = df_wid

ws['H'+str(head_line)] = c_h
ws['H'+str(head_line)].font = head_font
ws['H'+str(head_line)].alignment = default_align
ws.column_dimensions['H'].width = df_wid

ws['I'+str(head_line)] = c_i
ws['I'+str(head_line)].font = head_font
ws['I'+str(head_line)].alignment = default_align
ws.column_dimensions['I'].width = df_wid

ws['J'+str(head_line)] = c_j
ws['J'+str(head_line)].font = head_font
ws['J'+str(head_line)].alignment = default_align
ws.column_dimensions['J'].width = df_wid

ws['K'+str(head_line)] = c_k
ws['K'+str(head_line)].font = head_font
ws['K'+str(head_line)].alignment = default_align
ws.column_dimensions['K'].width = df_wid

ws['L'+str(head_line)] = c_l
ws['L'+str(head_line)].font = head_font
ws['L'+str(head_line)].alignment = default_align
ws.column_dimensions['L'].width = df_wid

ws['M'+str(head_line)] = c_m
ws['M'+str(head_line)].font = head_font
ws['M'+str(head_line)].alignment = default_align
ws.column_dimensions['M'].width = df_wid

ws['N'+str(head_line)] = c_n
ws['N'+str(head_line)].font = head_font
ws['N'+str(head_line)].alignment = default_align
ws.column_dimensions['N'].width = df_wid

ws['O'+str(head_line)] = c_o											# visual inspection
ws['O'+str(head_line)].font = head_font
ws['O'+str(head_line)].alignment = default_align
ws.column_dimensions['O'].width = 75

ws['P'+str(head_line)] = c_p
ws['P'+str(head_line)].font = head_font
ws['P'+str(head_line)].alignment = default_align
ws.column_dimensions['P'].width = 30

ws['Q'+str(head_line)] = c_q
ws['Q'+str(head_line)].font = head_font
ws['Q'+str(head_line)].alignment = default_align
ws.column_dimensions['Q'].width = 30


# write the names of the categories
# manufacturer types - A column
la = 2
while la <= (len(manufac)+1):
	for a in manufac:
		ws['A'+str(la)] = a
		ws['A'+str(la)].alignment = default_align
		la = la+1

# size types - B column
lb = 2
while lb <= (len(size)+1):
	for s in size:
		ws['B'+str(lb)] = s
		ws['B'+str(lb)].alignment = default_align
		lb = lb+1

# label type - C column
lc = 2
while lc <= (len(label)+1):
	for l in label:
		ws['C'+str(lc)] = l
		ws['C'+str(lc)].alignment = default_align
		lc = lc+1

# color types - D column
ld = 2
while ld <= (len(color)+1):
	for c in color:
		ws['D'+str(ld)] = c
		ld = ld+1

# write cable statistics data
i=2
with open('Data.csv','r') as dtrdr:
	while i <= (len(manufac)+1):
		lines = dtrdr.readlines()
		for line in lines:
			elem = line.split('\t')[0]
			ws['E'+str(i)] = elem
			ws['E'+str(i)].alignment = default_align
			elem = line.split('\t')[1]
			ws['F'+str(i)] = elem
			ws['F'+str(i)].alignment = default_align
			elem = line.split('\t')[2]
			ws['G'+str(i)] = elem
			ws['G'+str(i)].alignment = default_align
			elem = line.split('\t')[3]
			ws['H'+str(i)] = elem
			ws['H'+str(i)].alignment = default_align
			elem = line.split('\t')[4]
			ws['I'+str(i)] = elem
			ws['I'+str(i)].alignment = default_align
			elem = line.split('\t')[5]
			ws['J'+str(i)] = elem
			ws['J'+str(i)].alignment = default_align
			elem = line.split('\t')[6]
			ws['K'+str(i)] = elem
			ws['K'+str(i)].alignment = default_align
			elem = line.split('\t')[7]
			ws['L'+str(i)] = elem
			ws['L'+str(i)].alignment = default_align
			elem = line.split('\t')[8]
			ws['M'+str(i)] = elem
			ws['M'+str(i)].alignment = default_align
			elem = line.split('\t')[9]
			ws['N'+str(i)] = elem
			ws['N'+str(i)].alignment = default_align
			i=i+1

# write cable inspections data
i=2
with open('Inspection.csv','r') as insprdr:
	while i <= (len(manufac)+1):
		lines = insprdr.readlines()
		for line in lines:
			elem = line.split('\t')[0]
			ws['O'+str(i)] = elem
			ws['O'+str(i)].alignment = default_align
			elem = line.split('\t')[1]
			ws['P'+str(i)] = elem
			ws['P'+str(i)].alignment = default_align
			elem = line.split('\t')[2]
			ws['Q'+str(i)] = elem
			ws['Q'+str(i)].alignment = default_align
			i=i+1

# save the xls file
wb.save('Cable_tests.xls')

